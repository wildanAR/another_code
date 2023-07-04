from openpyxl import Workbook, load_workbook

"""MENGAKSES FILE YANG SUDAH ADA"""
# wb = load_workbook("contoh.xlsx") #load file excel
# ws = wb.active #Sheet yang sedang digunakan
# ws = wb['Sheet2'] #specific Sheet yang digunakan
# # print(ws)

# # print(ws["A1"].value) #nilai yang diambil pada cell A1

# # ws['A1'].value = 10 #change value

# # wb.save('contoh.xlsx') #save modify 

# print(wb.sheetnames) #melihat semua sheet yang tersedia

# wb.create_sheet("Sheet4") #membuat sheet baru
# wb.save('contoh.xlsx')

"""MEMBUAT WORKBOOK BARU ATAU FILE BARU"""
# wb = Workbook()
# ws = wb.active

# ws.title = "tes"

# #adding data
# ws.append(['wildan', 'awalusdfdin', 'rachman'])
# wb.save("biodata.xlsx")

"""Automate adding value in excel"""
# wb = load_workbook("biodata.xlsx")
# ws = wb['tes']

# ws.append(['wildan', 21, 'mahasiswa'])
# wb.save("biodata.xlsx")

"""ACCESS MULTIPLE CELL"""
# from openpyxl.utils import get_column_letter
# wb = load_workbook("biodata.xlsx")
# ws = wb.active

# for row in range(1,10):
#     for col in range(1,3):
#         char = get_column_letter(col)
#         # print(ws[char + str(row)].value)
#         ws[char + str(row)] = char + str(row) #edit value data

# wb.save('biodata.xlsx')

"""MERGING CELL"""
# wb = load_workbook("biodata.xlsx")
# ws = wb.active

# ws.merge_cells("A1:D1")

# ws.unmerge_cells("A1:D1")

# ws.insert_rows(2) 

# ws.delete_rows(2)

# ws.insert_cols(2)

# ws.delete_cols(2)
# wb.save('biodata.xlsx')

